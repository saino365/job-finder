#!/usr/bin/env python3
"""Update Excel with fixes for batch 4 - Defect sheet."""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook
    from openpyxl.styles import Font

def update_excel_fixes(defect_to_row_map, fixes, sheet_name="Defect", excel_path="JobFinder_TestCase.xlsx"):
    """Update Excel Defect sheet with fixes.
    
    Args:
        defect_to_row_map: dict mapping defect_id to row_number, e.g. {'D1': 2, 'D3': 4}
        fixes: list of fix summaries in same order as defect_to_row_map keys
    """
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    
    # Find Status and Fix Summary columns
    status_col_idx = None
    fix_summary_col_idx = None
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and 'Status' in str(header):
            status_col_idx = col
        if header and 'Fix Summary' in str(header):
            fix_summary_col_idx = col
    
    if not status_col_idx:
        print("Status column not found!")
        return False
    
    if not fix_summary_col_idx:
        # Add Fix Summary column after Status
        fix_summary_col_idx = status_col_idx + 1
        sheet.insert_cols(fix_summary_col_idx)
        sheet.cell(row=1, column=fix_summary_col_idx).value = "Fix Summary"
        sheet.cell(row=1, column=fix_summary_col_idx).font = Font(bold=True)
    
    # Update rows
    updated_count = 0
    defect_ids = list(defect_to_row_map.keys())
    for defect_id, fix_summary in zip(defect_ids, fixes):
        row_num = defect_to_row_map[defect_id]
        # Clear Status
        sheet.cell(row=row_num, column=status_col_idx).value = None
        # Add Fix Summary
        sheet.cell(row=row_num, column=fix_summary_col_idx).value = fix_summary
        updated_count += 1
        print(f"✓ Updated {defect_id} (Row {row_num})")
    
    wb.save(excel_path)
    print(f"\n✅ Updated {updated_count} defect(s) in Excel")
    return True

if __name__ == "__main__":
    # Defects from images with their row numbers: D1, D3, D4, D6, D7, D11, D12, D13, D18, D20, D21
    defect_to_row = {
        'D1': 2,
        'D3': 4,
        'D4': 5,
        'D6': 7,
        'D7': 8,
        'D11': 12,
        'D12': 13,
        'D13': 14,
        'D18': 19,
        'D20': 21,
        'D21': 22
    }
    
    fixes = [
        "D1: Email verification - After registration, user is redirected to main page. Email verification code is sent, but after relogin there's no option to enter verification code. Need to add verification code input option in login flow or profile settings for users with unverified emails.",
        "D3: Profile Education Details - Successfully saves education entry without filling required fields. Need to add validation to require all education fields (Level, Institution, Qualification, Field of Study) before allowing save.",
        "D4: Education Form - Form shows 'University' field which may be redundant with 'Institution' field. Need to review form structure and ensure consistent field naming (Institution vs University).",
        "D6: My Applications - Incorrect company name displayed. Shows '666-784-356/789/56/78/90/14' instead of actual company name 'TestCorp Solutions Sdn Bhd'. Need to fix company name lookup/display in applications table.",
        "D7: Job Listings Layout - Need to change job listings to grid layout. 'Apply now' button should be inside the job application box/card. Requires UI/UX redesign of job listings page.",
        "D11: Registration Validation - Username and email already exist in system, but user is able to proceed to create new account with same email and username. Need to add validation to check for existing username/email and show error messages: 'This email address is already registered' and 'This username is already registered'.",
        "D12: Company Registration - Test using same email that account exists, no error message after clicking 'Create Account' for Register as employer. Need to add duplicate email/username validation for company registration form. Should show: 'This email address is already registered' and 'This username is already registered'.",
        "D13: Profile Score - Profile score shows 76% but no pending actions are displayed. Need to ensure pending actions section shows incomplete profile fields that need to be filled to reach 100%.",
        "D18: Profile Interest Section - No 'Interest' section appears in the profile after login, even though it was filled during account creation. Need to ensure Interest data from registration is saved and displayed in profile page.",
        "D20: Profile Event Section - No 'Event' section appears in the profile after login. Need to ensure Event data from registration is saved and displayed in profile page.",
        "D21: Recommended Jobs - Manufacturing exists in job list and is set as preferred industry in profile filter, but no results show in 'Recommended For You' section. Need to fix recommendation algorithm to properly filter and display jobs based on user's preferred industries."
    ]
    
    update_excel_fixes(defect_to_row, fixes)
