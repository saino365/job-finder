#!/usr/bin/env python3
"""
Extract images from Excel file for analysis.
"""
import os
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl pillow --quiet")
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

def extract_images_from_excel(excel_path, output_dir="extracted_images"):
    """Extract all images from an Excel file."""
    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        return
    
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    
    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    
    image_count = 0
    
    # Extract images from all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")
        
        # Check if sheet has images
        if hasattr(sheet, '_images') and sheet._images:
            for idx, img in enumerate(sheet._images):
                image_count += 1
                # Get image data
                img_data = img._data()
                
                # Determine file extension from image format
                ext = 'png'  # default
                if hasattr(img, 'format') and img.format:
                    ext = img.format.lower()
                elif img_data.startswith(b'\x89PNG'):
                    ext = 'png'
                elif img_data.startswith(b'\xff\xd8'):
                    ext = 'jpg'
                elif img_data.startswith(b'GIF'):
                    ext = 'gif'
                
                # Save image
                filename = f"{sheet_name}_image_{idx+1}.{ext}"
                filepath = output_dir / filename
                
                with open(filepath, 'wb') as f:
                    f.write(img_data)
                
                print(f"  ✓ Extracted: {filename} ({len(img_data)} bytes)")
        
        # Also check for embedded images in cells (less common)
        # This would require checking cell comments, hyperlinks, etc.
    
    print(f"\n✅ Extracted {image_count} image(s) to '{output_dir}/' directory")
    print(f"\nYou can now paste these images in chat for analysis!")
    
    return output_dir

if __name__ == "__main__":
    excel_file = "JobFinder_TestCase.xlsx"
    extract_images_from_excel(excel_file)
