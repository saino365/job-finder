#!/usr/bin/env python3
"""Update Excel with fixes for batch 5 - Defect sheet."""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook
    from openpyxl.styles import Font

def update_excel_fixes(defect_to_row_map, fixes, sheet_name="Defect", excel_path="JobFinder_TestCase.xlsx"):
    """Update Excel Defect sheet with fixes.
    
    Args:
        defect_to_row_map: dict mapping defect_id to row_number, e.g. {'D22': 23, 'D23': 24}
        fixes: list of fix summaries in same order as defect_to_row_map keys
    """
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    
    # Find Status and Fix Summary columns
    status_col_idx = None
    fix_summary_col_idx = None
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and 'Status' in str(header):
            status_col_idx = col
        if header and 'Fix Summary' in str(header):
            fix_summary_col_idx = col
    
    if not status_col_idx:
        print("Status column not found!")
        return False
    
    if not fix_summary_col_idx:
        # Add Fix Summary column after Status
        fix_summary_col_idx = status_col_idx + 1
        sheet.insert_cols(fix_summary_col_idx)
        sheet.cell(row=1, column=fix_summary_col_idx).value = "Fix Summary"
        sheet.cell(row=1, column=fix_summary_col_idx).font = Font(bold=True)
    
    # Update rows
    updated_count = 0
    defect_ids = list(defect_to_row_map.keys())
    for defect_id, fix_summary in zip(defect_ids, fixes):
        row_num = defect_to_row_map[defect_id]
        # Clear Status
        sheet.cell(row=row_num, column=status_col_idx).value = None
        # Add Fix Summary
        sheet.cell(row=row_num, column=fix_summary_col_idx).value = fix_summary
        updated_count += 1
        print(f"✓ Updated {defect_id} (Row {row_num})")
    
    wb.save(excel_path)
    print(f"\n✅ Updated {updated_count} defect(s) in Excel")
    return True

if __name__ == "__main__":
    # Defects from images with their row numbers: D22, D23, D24, D25, D26, D27, D29, D31, D33, D34, D35
    defect_to_row = {
        'D22': 23,
        'D23': 24,
        'D24': 25,
        'D25': 26,
        'D26': 27,
        'D27': 28,
        'D29': 30,
        'D31': 32,
        'D33': 34,
        'D34': 35,
        'D35': 36
    }
    
    fixes = [
        "D22: Certificate Date Validation - Certificate issue date must not be set in the future. Need to add date validation to ensure acquired/issue date cannot be in the future. Validation should work across all browsers and devices (including incognito and mobile).",
        "D23: Work Experience Date Fields - Start Date, End Date, and Ongoing(Present) option can be added during registration but are missing from work experience section in profile after login. Need to add Start Date, End Date fields and Ongoing(Present) checkbox/button to work experience edit form. Also fix issue where End Date cannot be selected - ensure date picker works correctly.",
        "D24: Application Validity Extension - After changing 'Extend Application Validity' to specified days and clicking Extend, no popup notification appears and the validity days area is not updated. Need to add success notification and update the validity date display immediately after extension.",
        "D25: Apply Now Button - Apply Now button redirects to job details page instead of applying for the job. Need to fix button functionality to properly submit application or redirect to application form, not just show job details.",
        "D26: Recommendation System - Duration Filter - Jobs matching the preferred internship duration exist in the system but are not appearing in 'Recommended For You' section. Need to fix recommendation algorithm to properly filter and display jobs based on user's preferred internship duration.",
        "D27: Recommendation System - Salary Filter - Salary filter in Internship Details is set but recommended jobs show different salary ranges, indicating filter logic may not be applied correctly. Also, when salary range exists, no recommended jobs are shown. Need to fix recommendation algorithm to properly filter jobs by salary range and display matching results.",
        "D29: GPA Field Validation - System does not save alphabetic input in GPA field. Need to add validation to accept only numeric input (with optional decimal point) for GPA field. Should reject alphabetic characters.",
        "D31: Phone Number Field Validation - Phone number field incorrectly allows non-numeric characters. Field should only allow numeric digits and optional plus sign (+) at the start. Need to add proper validation to prevent saving invalid characters like '@' and alphabetic characters.",
        "D33: Industry Filter Update - Filter for 'Technology' on Job page needs to be updated to 'Information Technology'. Industry for job posts also needs to reflect this change. Need to update filter options and ensure consistency across job listings and filters.",
        "D34: Internship Start Date Filter - Filter by 'This Month' for 'Internship Start Date' shows 0 jobs found even when jobs with matching dates exist. Need to fix date filter logic to properly match jobs with start dates within the current month.",
        "D35: Mobile Horizontal Scrolling - Horizontal (left & right) scrolling is not functional on mobile view, both before and after login. Need to fix CSS/layout to enable proper horizontal scrolling on mobile devices for all pages."
    ]
    
    update_excel_fixes(defect_to_row, fixes)
