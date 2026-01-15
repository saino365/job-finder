#!/usr/bin/env python3
"""Update Excel with fixes for batch 7 - Defect sheet."""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook
    from openpyxl.styles import Font

def update_excel_fixes(defect_to_row_map, fixes, sheet_name="Defect", excel_path="JobFinder_TestCase.xlsx"):
    """Update Excel Defect sheet with fixes.
    
    Args:
        defect_to_row_map: dict mapping defect_id to row_number, e.g. {'D91': 92, 'D92': 93}
        fixes: list of fix summaries in same order as defect_to_row_map keys
    """
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    
    # Find Status and Fix Summary columns
    status_col_idx = None
    fix_summary_col_idx = None
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and 'Status' in str(header):
            status_col_idx = col
        if header and 'Fix Summary' in str(header):
            fix_summary_col_idx = col
    
    if not status_col_idx:
        print("Status column not found!")
        return False
    
    if not fix_summary_col_idx:
        # Add Fix Summary column after Status
        fix_summary_col_idx = status_col_idx + 1
        sheet.insert_cols(fix_summary_col_idx)
        sheet.cell(row=1, column=fix_summary_col_idx).value = "Fix Summary"
        sheet.cell(row=1, column=fix_summary_col_idx).font = Font(bold=True)
    
    # Update rows
    updated_count = 0
    defect_ids = list(defect_to_row_map.keys())
    for defect_id, fix_summary in zip(defect_ids, fixes):
        row_num = defect_to_row_map[defect_id]
        # Clear Status
        sheet.cell(row=row_num, column=status_col_idx).value = None
        # Add Fix Summary
        sheet.cell(row=row_num, column=fix_summary_col_idx).value = fix_summary
        updated_count += 1
        print(f"✓ Updated {defect_id} (Row {row_num})")
    
    wb.save(excel_path)
    print(f"\n✅ Updated {updated_count} defect(s) in Excel")
    return True

if __name__ == "__main__":
    # Defects from images with their row numbers: D91-D109
    defect_to_row = {
        'D91': 92,
        'D92': 93,
        'D93': 94,
        'D94': 95,
        'D95': 96,
        'D96': 97,
        'D97': 98,
        'D98': 99,
        'D99': 100,
        'D100': 101,
        'D101': 102,
        'D102': 103,
        'D103': 104,
        'D104': 105,
        'D105': 106,
        'D106': 107,
        'D107': 108,
        'D108': 109,
        'D109': 110
    }
    
    fixes = [
        "D91: Job Creation Fields - Requirements, Responsibilities, and Onboarding Materials fields are not available during job creation but appear in job posting. Need to add these fields to the job creation form so they can be filled during creation.",
        "D92: File Upload Notification - Files can be uploaded in create job listing, but no notification message appears and files are not viewable after job is created. Need to add upload confirmation message and ensure uploaded files are properly saved and displayed after job creation.",
        "D93: Posted vs Published Date - Posted and Published dates function the same after admin approval. Need to clarify the difference between Posted and Published dates, or fix Published date to reflect actual publication time. Currently shows wrong Published date (07/11/2025 when job was viewable on 06/11/2025).",
        "D94: Location Filter - Search by location filter shows no jobs found even when location exists in the filter list. Need to fix location filter logic to properly match and display jobs based on selected location.",
        "D95: Project Date Validation - System allows job postings with project start and end dates set in previous months. Need to add validation to restrict dates to current or future months only.",
        "D96: Internship Date Validation - System allows job postings with internship start and end dates set in previous months. Need to add validation to restrict dates to current or future months only.",
        "D97: Internship Date Display - Internship Start and End dates are added in job posting form but not visible in the job listing view. Need to ensure these dates are properly saved and displayed in job listing details.",
        "D98: Instant Publish Function - 'Instant Publish upon Approval' function automatically publishes job listing immediately after admin approval, regardless of checkbox state. Also missing 'Publish' button in Job Management page. Need to fix checkbox functionality and add manual publish option.",
        "D99: Date Logic Validation - System allows job listing submission with start date in the future compared to end date (e.g., Start: 01/2026, End: 12/2025). Need to add validation to ensure start date is not later than end date.",
        "D100: Publish Date Validation - Can select 'Specific date/time upon approval' with past date. Need to add validation to restrict this field to current or future dates only.",
        "D101: Location Field Validation - System allows saving short/incomplete location entries (e.g., 'cl') when creating city/state in job listing. Should require full location names matching the location filter options. Need to add validation to ensure complete location names are entered.",
        "D102: Salary Validation - System allows setting minimum salary higher than maximum salary. Need to add validation to ensure minimum salary is less than or equal to maximum salary.",
        "D103: Salary Range Display - Minimum salary (RM 40,000) is higher than maximum salary (RM 7,000) in job details. This is a display/validation issue. Need to fix salary validation and ensure proper range display.",
        "D104: Featured Companies Filter - Filter by salary shows positions available but no results under featured companies. Need to fix filter logic to properly show featured companies when jobs match the salary filter criteria.",
        "D105: Resume File Format - System allows image format files (PNG) to be uploaded as resume. Should only allow document formats (PDF, DOC, DOCX). Need to add file type validation to restrict resume uploads to appropriate document formats.",
        "D106: Certificate File Validation - No error message when uploading invalid file types (.xsl, image formats) for certificates. Need to add file type validation and show appropriate error messages for invalid certificate file formats.",
        "D107: Reject Application Button - Reject candidate button doesn't function in the reject application modal. Need to fix button functionality to properly process application rejection.",
        "D108: Company Email Notifications - Company doesn't receive email notifications for 'job listing approved' or 'New application from candidate' events, even though notifications appear in UI. Need to implement email notification system for these events.",
        "D109: Session Management - When browser is closed without logout and reopened, system shows 'user is unknown' instead of redirecting to login page. Need to fix session management to properly handle expired sessions and redirect to login page."
    ]
    
    update_excel_fixes(defect_to_row, fixes)
