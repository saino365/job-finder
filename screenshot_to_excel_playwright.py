#!/usr/bin/env python3
"""
Take a screenshot of a webpage using Playwright and add it to an Excel file row.
Playwright is easier to set up than Selenium - it auto-downloads browsers.
"""
import os
import sys
from pathlib import Path
from datetime import datetime

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Installing playwright...")
    os.system(f"{sys.executable} -m pip install playwright --quiet")
    print("Installing browser... (this may take a minute)")
    os.system(f"{sys.executable} -m playwright install chromium")
    from playwright.sync_api import sync_playwright

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl pillow --quiet")
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

def take_screenshot(url, output_path="screenshot.png", wait_seconds=3, full_page=True):
    """
    Take a screenshot of a webpage using Playwright.
    
    Args:
        url: URL to screenshot
        output_path: Path to save screenshot
        wait_seconds: Seconds to wait for page to load
        full_page: If True, capture full page
    """
    print(f"Opening browser for: {url}")
    
    with sync_playwright() as p:
        # Launch browser
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(viewport={'width': 1920, 'height': 1080})
        page = context.new_page()
        
        try:
            # Navigate to URL
            print(f"Loading page...")
            page.goto(url, wait_until='networkidle', timeout=30000)
            
            # Wait additional time if needed
            if wait_seconds > 0:
                print(f"Waiting {wait_seconds} seconds...")
                page.wait_for_timeout(wait_seconds * 1000)
            
            # Take screenshot
            screenshot_path = Path(output_path)
            if full_page:
                page.screenshot(path=str(screenshot_path), full_page=True)
            else:
                page.screenshot(path=str(screenshot_path))
            
            print(f"✅ Screenshot saved: {screenshot_path}")
            return str(screenshot_path)
        
        except Exception as e:
            print(f"Error taking screenshot: {e}")
            return None
        
        finally:
            browser.close()

def add_image_to_excel(excel_path, sheet_name, row, col, image_path, image_width=400, image_height=300):
    """
    Add an image to an Excel file at a specific cell.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of the sheet
        row: Row number (1-indexed)
        col: Column letter (e.g., 'A', 'B', 'C')
        image_path: Path to image file
        image_width: Width of image in pixels
        image_height: Height of image in pixels
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}")
        return False
    
    image_path = Path(image_path)
    if not image_path.exists():
        print(f"Error: Image file not found: {image_path}")
        return False
    
    print(f"Opening Excel file: {excel_path}")
    wb = load_workbook(excel_path)
    
    # Get or create sheet
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Creating it...")
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    
    # Load and resize image
    img = Image(str(image_path))
    img.width = image_width
    img.height = image_height
    
    # Calculate cell position
    cell = f"{col}{row}"
    
    # Anchor image to cell
    img.anchor = cell
    
    # Add image to worksheet
    ws.add_image(img)
    
    # Adjust row height to fit image
    ws.row_dimensions[row].height = image_height * 0.75  # Convert pixels to Excel points
    
    # Save workbook
    wb.save(excel_path)
    print(f"✅ Image added to {sheet_name}!{cell} in {excel_path}")
    
    return True

def screenshot_and_add_to_excel(url, excel_path, sheet_name, row, col='A', 
                                 wait_seconds=3, image_width=400, image_height=300):
    """
    Complete workflow: Take screenshot and add to Excel.
    
    Args:
        url: URL to screenshot
        excel_path: Path to Excel file
        sheet_name: Name of sheet
        row: Row number (1-indexed)
        col: Column letter (default 'A')
        wait_seconds: Seconds to wait for page load
        image_width: Image width in pixels
        image_height: Image height in pixels
    """
    # Generate temporary screenshot filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    screenshot_path = f"screenshot_{timestamp}.png"
    
    # Take screenshot
    screenshot_file = take_screenshot(url, screenshot_path, wait_seconds)
    
    if not screenshot_file:
        print("❌ Failed to take screenshot")
        return False
    
    # Add to Excel
    success = add_image_to_excel(excel_path, sheet_name, row, col, screenshot_file, 
                                  image_width, image_height)
    
    return success

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Take webpage screenshot and add to Excel (using Playwright)')
    parser.add_argument('url', help='URL to screenshot')
    parser.add_argument('--excel', '-e', default='JobFinder_TestCase.xlsx', 
                       help='Excel file path (default: JobFinder_TestCase.xlsx)')
    parser.add_argument('--sheet', '-s', default='Screenshots', 
                       help='Sheet name (default: Screenshots)')
    parser.add_argument('--row', '-r', type=int, default=1, 
                       help='Row number (default: 1)')
    parser.add_argument('--col', '-c', default='A', 
                       help='Column letter (default: A)')
    parser.add_argument('--wait', '-w', type=int, default=3, 
                       help='Wait seconds for page load (default: 3)')
    parser.add_argument('--width', type=int, default=400, 
                       help='Image width in pixels (default: 400)')
    parser.add_argument('--height', type=int, default=300, 
                       help='Image height in pixels (default: 300)')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("Screenshot to Excel Tool (Playwright)")
    print("=" * 60)
    print(f"URL: {args.url}")
    print(f"Excel: {args.excel}")
    print(f"Sheet: {args.sheet}, Row: {args.row}, Col: {args.col}")
    print("=" * 60)
    
    screenshot_and_add_to_excel(
        args.url,
        args.excel,
        args.sheet,
        args.row,
        args.col,
        args.wait,
        args.width,
        args.height
    )
