#!/usr/bin/env python3
"""
Get full details of failed test cases.
"""
import sys
import os
from pathlib import Path
import json

try:
    from openpyxl import load_workbook
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook

def get_test_details(row_numbers, sheet_name="Registration.Login", excel_path="JobFinder_TestCase.xlsx"):
    """Get full details of specific test cases."""
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[sheet_name]
    
    # Get headers
    headers = []
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        headers.append(header)
    
    tests = []
    for row_num in row_numbers:
        test_data = {
            'row_number': row_num,
            'headers': headers
        }
        for col_idx, header in enumerate(headers, 1):
            cell_value = sheet.cell(row=row_num, column=col_idx).value
            test_data[header] = cell_value
        
        tests.append(test_data)
    
    return tests

if __name__ == "__main__":
    # First 10 failed test rows: 4, 5, 11, 14, 15, 16, 17, 20, 21, 22
    row_numbers = [4, 5, 11, 14, 15, 16, 17, 20, 21, 22]
    tests = get_test_details(row_numbers)
    
    print("="*60)
    print("DETAILED TEST CASE INFORMATION")
    print("="*60)
    
    for idx, test in enumerate(tests, 1):
        print(f"\n{'='*60}")
        print(f"TEST {idx}: Row {test['row_number']}")
        print(f"{'='*60}")
        print(f"Test Case No: {test.get('Test Case No', 'N/A')}")
        print(f"Test Cases: {test.get('Test Cases', 'N/A')}")
        print(f"Test Scenario: {test.get('Test Scenario', 'N/A')}")
        print(f"Test Steps:\n{test.get('Test Steps', 'N/A')}")
        print(f"Expected Results:\n{test.get('Expected Results', 'N/A')}")
        print(f"Actual Result:\n{test.get('Actual Result', 'N/A')}")
        print(f"Status: {test.get('Status', 'N/A')}")
        print(f"Defect: {test.get('Defect', 'N/A')}")
        print(f"Test Data: {test.get('Test Data', 'N/A')}")
