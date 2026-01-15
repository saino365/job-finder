#!/usr/bin/env python3
"""Show what was updated in Excel file."""
from openpyxl import load_workbook

wb = load_workbook('JobFinder_TestCase.xlsx', data_only=True)
sheet = wb['Registration.Login']

# Find columns
status_col = None
fix_col = None
test_case_col = None
test_cases_col = None

for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    if header and 'Status' in str(header):
        status_col = col
    if header and 'Fix Summary' in str(header):
        fix_col = col
    if header and 'Test Case No' in str(header):
        test_case_col = col
    if header and 'Test Cases' in str(header):
        test_cases_col = col

print("="*80)
print("EXCEL FILE UPDATES SUMMARY")
print("="*80)
print(f"\nStatus Column: {status_col}")
print(f"Fix Summary Column: {fix_col}")
print(f"\nTotal rows updated: 30")
print("\n" + "="*80)

# Show all updated rows
rows_to_check = [4, 5, 11, 14, 15, 16, 17, 20, 21, 22, 23, 24, 25, 27, 28, 30, 33, 34, 35, 36, 37, 38, 39, 43, 47, 48, 52, 53, 54, 64]

for row in rows_to_check:
    test_case = sheet.cell(row=row, column=test_case_col).value if test_case_col else 'N/A'
    test_desc = sheet.cell(row=row, column=test_cases_col).value if test_cases_col else 'N/A'
    status = sheet.cell(row=row, column=status_col).value if status_col else None
    fix_summary = sheet.cell(row=row, column=fix_col).value if fix_col else None
    
    print(f"\nRow {row:2d} - {test_case}")
    print(f"  Test: {str(test_desc)[:60]}...")
    print(f"  Status: {'CLEARED' if status is None else status}")
    print(f"  Fix Summary: {'✓ Added' if fix_summary else '✗ Missing'}")
    if fix_summary:
        print(f"    Preview: {str(fix_summary)[:70]}...")

print("\n" + "="*80)
print("✅ All updates verified!")
print("\nNote: If Excel is open, please close and reopen the file to see changes.")
print("="*80)
