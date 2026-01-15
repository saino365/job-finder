#!/usr/bin/env python3
"""
Analyze the Excel file structure to understand how sheets work,
what data is in rows/columns, and how images relate to the data.
"""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

def analyze_excel_structure(excel_path="JobFinder_TestCase.xlsx"):
    """Analyze the structure of the Excel file."""
    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        return None
    
    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    
    structure = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Analyzing sheet: {sheet_name}")
        print(f"{'='*60}")
        
        # Get sheet dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        print(f"Dimensions: {max_row} rows x {max_col} columns")
        
        # Get column headers (first row)
        headers = []
        if max_row > 0:
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(cell_value if cell_value else f"Column{col}")
        
        print(f"Headers: {headers}")
        
        # Count images
        image_count = len(sheet._images) if hasattr(sheet, '_images') else 0
        print(f"Images: {image_count}")
        
        # Sample first few rows of data
        sample_rows = []
        for row_idx in range(1, min(6, max_row + 1)):  # First 5 rows
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_data[header] = str(cell_value)[:100]  # Truncate long values
            if row_data:
                sample_rows.append(row_data)
        
        # Get all non-empty rows summary
        data_rows = []
        for row_idx in range(2, max_row + 1):  # Skip header
            row_data = {}
            has_data = False
            for col_idx, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_data[header] = cell_value
                    has_data = True
            if has_data:
                data_rows.append(row_data)
        
        structure[sheet_name] = {
            'dimensions': (max_row, max_col),
            'headers': headers,
            'image_count': image_count,
            'data_row_count': len(data_rows),
            'sample_rows': sample_rows[:3],  # First 3 sample rows
            'all_data': data_rows[:10]  # First 10 data rows for analysis
        }
        
        print(f"\nData rows (non-empty): {len(data_rows)}")
        print(f"\nSample data (first 3 rows):")
        for idx, row in enumerate(sample_rows[:3], 1):
            print(f"  Row {idx}: {row}")
    
    return structure

def generate_structure_report(structure, output_file="excel_structure_report.md"):
    """Generate a markdown report of the Excel structure."""
    report_lines = [
        "# Excel File Structure Analysis",
        "",
        "## Overview",
        ""
    ]
    
    total_sheets = len(structure)
    total_images = sum(s['image_count'] for s in structure.values())
    total_data_rows = sum(s['data_row_count'] for s in structure.values())
    
    report_lines.extend([
        f"- **Total Sheets:** {total_sheets}",
        f"- **Total Images:** {total_images}",
        f"- **Total Data Rows:** {total_data_rows}",
        "",
        "---",
        ""
    ])
    
    for sheet_name, sheet_data in structure.items():
        report_lines.extend([
            f"## Sheet: {sheet_name}",
            "",
            f"**Dimensions:** {sheet_data['dimensions'][0]} rows × {sheet_data['dimensions'][1]} columns",
            f"**Images:** {sheet_data['image_count']}",
            f"**Data Rows:** {sheet_data['data_row_count']}",
            "",
            "### Column Headers",
            ""
        ])
        
        for idx, header in enumerate(sheet_data['headers'], 1):
            report_lines.append(f"{idx}. {header}")
        
        report_lines.append("")
        report_lines.append("### Sample Data")
        report_lines.append("")
        
        for idx, row in enumerate(sheet_data['sample_rows'], 1):
            report_lines.append(f"#### Row {idx + 1} (after header):")
            for key, value in row.items():
                # Truncate long values
                display_value = str(value)[:200] + "..." if len(str(value)) > 200 else str(value)
                report_lines.append(f"- **{key}:** {display_value}")
            report_lines.append("")
        
        report_lines.append("---")
        report_lines.append("")
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    
    print(f"\n✅ Structure report saved: {output_file}")
    return output_file

if __name__ == "__main__":
    print("="*60)
    print("Excel Structure Analyzer")
    print("="*60)
    
    structure = analyze_excel_structure()
    
    if structure:
        generate_structure_report(structure)
        print("\n" + "="*60)
        print("✅ Analysis Complete!")
        print("="*60)
