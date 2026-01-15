#!/usr/bin/env python3
"""Update Excel with fixes for batch 6 - Defect sheet."""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook
    from openpyxl.styles import Font

def update_excel_fixes(defect_to_row_map, fixes, sheet_name="Defect", excel_path="JobFinder_TestCase.xlsx"):
    """Update Excel Defect sheet with fixes.
    
    Args:
        defect_to_row_map: dict mapping defect_id to row_number, e.g. {'D36': 37, 'D39': 40}
        fixes: list of fix summaries in same order as defect_to_row_map keys
    """
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    
    # Find Status and Fix Summary columns
    status_col_idx = None
    fix_summary_col_idx = None
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and 'Status' in str(header):
            status_col_idx = col
        if header and 'Fix Summary' in str(header):
            fix_summary_col_idx = col
    
    if not status_col_idx:
        print("Status column not found!")
        return False
    
    if not fix_summary_col_idx:
        # Add Fix Summary column after Status
        fix_summary_col_idx = status_col_idx + 1
        sheet.insert_cols(fix_summary_col_idx)
        sheet.cell(row=1, column=fix_summary_col_idx).value = "Fix Summary"
        sheet.cell(row=1, column=fix_summary_col_idx).font = Font(bold=True)
    
    # Update rows
    updated_count = 0
    defect_ids = list(defect_to_row_map.keys())
    for defect_id, fix_summary in zip(defect_ids, fixes):
        row_num = defect_to_row_map[defect_id]
        # Clear Status
        sheet.cell(row=row_num, column=status_col_idx).value = None
        # Add Fix Summary
        sheet.cell(row=row_num, column=fix_summary_col_idx).value = fix_summary
        updated_count += 1
        print(f"✓ Updated {defect_id} (Row {row_num})")
    
    wb.save(excel_path)
    print(f"\n✅ Updated {updated_count} defect(s) in Excel")
    return True

if __name__ == "__main__":
    # Defects from images with their row numbers: D36, D39, D42, D43, D44, D75, D76, D77, D78, D79, D80, D81, D82, D83, D84, D85, D86, D87, D88, D89, D90
    defect_to_row = {
        'D36': 37,
        'D39': 40,
        'D42': 43,
        'D43': 44,
        'D44': 45,
        'D75': 76,
        'D76': 77,
        'D77': 78,
        'D78': 79,
        'D79': 80,
        'D80': 81,
        'D81': 82,
        'D82': 83,
        'D83': 84,
        'D84': 85,
        'D85': 86,
        'D86': 87,
        'D87': 88,
        'D88': 89,
        'D89': 90,
        'D90': 91
    }
    
    fixes = [
        "D36: Registration Page Layout - Layout doesn't work properly in register page when opened in landscape mode from mobile. Need to fix responsive CSS/layout to properly handle landscape orientation on mobile devices.",
        "D39: Application Status - Incorrect company name displayed in 'Reject status' applications. Shows incorrect name instead of actual company name. Need to fix company name lookup/display in rejected applications table.",
        "D42: Username Validation - Username should not allow only special characters. Must contain at least 3 alphabetic characters. Need to add validation to require minimum 3 alphabetic characters in username field.",
        "D43: Username Validation - Username should not allow only numeric characters. Must contain at least 3 alphabetic characters. Need to add validation to require minimum 3 alphabetic characters in username field (cannot be numeric-only).",
        "D44: Name Field Validation - First Name, Middle Name, and Last Name should not allow only special characters. Must contain at least 3 alphabetic characters. Need to add validation to require minimum 3 alphabetic characters in all name fields.",
        "D75: Company Verification Popup - 'Company Verification In Progress' popup keeps appearing each time user clicks on 'Profile', 'Job Management', 'Applications', 'Employees', 'Universities & Programmes', and 'Create jobs'. Need to add logic to show popup only once per session or add 'Don't show again' option.",
        "D76: Employee Page - Employee page is empty with no candidate status after applying for job position. Need to ensure that when candidates apply for jobs, their status appears correctly in the company's employee page.",
        "D77: Email Notification - Company does not receive email notification when candidate applies for job. Need to implement email notification system to send emails to company when candidates apply for positions.",
        "D78: Admin Dashboard - Admin not able to view details for application and employment status. Need to add functionality to allow admin to view detailed information for applications and employments from the dashboard status sections.",
        "D79: Interns Information - No interns info displayed in candidate/main page. Need to ensure intern candidates are properly displayed in the main candidates page with their information visible.",
        "D80: Application Location - In the apply for internship under application, no setting to add candidate location in the profile. Need to add location field in profile settings that can be used during application process.",
        "D81: Profile Completeness Check - Candidate profile is updated, but during apply for internship under application, there's error message 'Incomplete Profile' even though profile is complete. Need to fix profile completeness validation logic to accurately reflect profile status.",
        "D82: Company Management Status Filter - Status 'Pending' or 'Rejected' companies still show as 'Approved' in all tabs of Company Management. Need to fix status filtering logic to correctly display companies based on their actual status in each tab.",
        "D83: Application Notification - Didn't receive notification in the system once the company responds to job application. Need to implement notification system to alert candidates when companies respond to their applications.",
        "D84: Application Status Sync - Candidate has accepted the job offer, but the status in company view still shows 'Pending acceptance'. Need to fix status synchronization between candidate and company views when candidate accepts/rejects offers.",
        "D85: Employee Termination Display - Unknown name displayed in termination list under employees (shows numerical ID instead of name). Need to fix employee name lookup/display in termination list to show actual employee names.",
        "D86: Job Management Delete - Delete button doesn't function under Job Management. Need to implement delete functionality for job listings in the Job Management section.",
        "D87: Edit Job Listing Cancel - 'Cancel' button doesn't function in edit job listing page. Need to implement cancel functionality to allow users to exit edit mode without saving changes.",
        "D88: Edit Job Listing Data Loss - Missing data in 'Edit Job Listing' after clicking submit or draft. Data entered in form is not being saved properly. Need to fix form submission and draft saving to preserve all entered data.",
        "D89: Job Details Display - Missing data for 'Profession', 'Location', and 'Salary Range' after viewing job details. Data is saved but not displayed correctly in job details view. Need to fix data retrieval and display in job details page.",
        "D90: Monitoring Job Details - Missing 'Profession' info in Monitoring after clicking view. Need to ensure all job details including profession are properly displayed in the monitoring dashboard view."
    ]
    
    update_excel_fixes(defect_to_row, fixes)
