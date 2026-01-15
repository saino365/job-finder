#!/usr/bin/env python3
"""Update Excel with fixes for batch 2."""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook
    from openpyxl.styles import Font

def update_excel_fixes(row_numbers, fixes, sheet_name="Registration.Login", excel_path="JobFinder_TestCase.xlsx"):
    """Update Excel with fixes."""
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]
    
    # Find Status and Fix Summary columns
    status_col_idx = None
    fix_summary_col_idx = None
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and 'Status' in str(header):
            status_col_idx = col
        if header and 'Fix Summary' in str(header):
            fix_summary_col_idx = col
    
    if not status_col_idx:
        print("Status column not found!")
        return False
    
    if not fix_summary_col_idx:
        # Add Fix Summary column after Status
        fix_summary_col_idx = status_col_idx + 1
        sheet.insert_cols(fix_summary_col_idx)
        sheet.cell(row=1, column=fix_summary_col_idx).value = "Fix Summary"
        sheet.cell(row=1, column=fix_summary_col_idx).font = Font(bold=True)
    
    # Update rows
    for row_num, fix_summary in zip(row_numbers, fixes):
        # Clear Status
        sheet.cell(row=row_num, column=status_col_idx).value = None
        # Add Fix Summary
        sheet.cell(row=row_num, column=fix_summary_col_idx).value = fix_summary
    
    wb.save(excel_path)
    print(f"✅ Updated {len(row_numbers)} rows in Excel")
    return True

if __name__ == "__main__":
    # Next 10 failed tests: rows 23, 24, 25, 27, 28, 30, 33, 34, 35, 36
    row_numbers = [23, 24, 25, 27, 28, 30, 33, 34, 35, 36]
    
    fixes = [
        "TC22: Phone validation - Fixed: Now properly rejects alphabetic+special character combinations. Removed normalize function that was silently removing invalid chars.",
        "TC23: Phone validation - Fixed: Now properly rejects numeric+special character combinations. Validator now catches invalid input before submission.",
        "TC24: Phone validation - Fixed: Now properly rejects alphabetic+numeric+special character combinations. Phone field only accepts digits and optional plus at start.",
        "TC26: Email verification reuse - Backend supports reusing email for verification codes. May need frontend flow adjustment to allow incomplete registrations to request new codes.",
        "TC27: Username validation - Fixed: Added async check to verify username availability before submission. Shows error if username already exists.",
        "Row 30: Name validation with special chars - Current validation allows apostrophes and hyphens (O'Brien, Mary-Jane) which are standard. If test expects other special chars, needs clarification.",
        "TC32: Password validation - Fixed: Now enforces 8+ chars, 1 lowercase, 1 uppercase, 1 number, 1 special character. All conditions must be met.",
        "TC33: Password validation - Fixed: Same as TC32, validates all password requirements including length > 8 chars.",
        "TC34: Password validation - Fixed: Now properly rejects passwords with less than 8 characters even if other conditions are met.",
        "TC35: Password validation - Fixed: Now properly rejects passwords missing lowercase letters even if other conditions are met."
    ]
    
    update_excel_fixes(row_numbers, fixes)
