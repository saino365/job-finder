#!/usr/bin/env python3
"""Get full details of next 10 failed test cases."""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook

def get_test_details(row_numbers, sheet_name="Registration.Login", excel_path="JobFinder_TestCase.xlsx"):
    """Get full details of specific test cases."""
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[sheet_name]
    
    headers = []
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        headers.append(header)
    
    tests = []
    for row_num in row_numbers:
        test_data = {'row_number': row_num}
        for col_idx, header in enumerate(headers, 1):
            cell_value = sheet.cell(row=row_num, column=col_idx).value
            test_data[header] = cell_value
        tests.append(test_data)
    
    return tests

if __name__ == "__main__":
    # Get row numbers from user input or use default
    import sys
    if len(sys.argv) > 1:
        row_numbers = [int(x) for x in sys.argv[1:]]
    else:
        # Default: next 10 after already fixed ones
        row_numbers = []  # Will be determined by the calling script
    
    tests = get_test_details(row_numbers)
    
    for idx, test in enumerate(tests, 1):
        print(f"\n{'='*60}")
        print(f"TEST {idx}: Row {test['row_number']} - {test.get('Test Case No', 'N/A')}")
        print(f"{'='*60}")
        print(f"Test Cases: {test.get('Test Cases', 'N/A')}")
        print(f"Test Steps:\n{test.get('Test Steps', 'N/A')}")
        print(f"Expected Results:\n{test.get('Expected Results', 'N/A')}")
        print(f"Actual Result:\n{test.get('Actual Result', 'N/A')}")
        print(f"Status: {test.get('Status', 'N/A')}")
        print(f"Defect: {test.get('Defect', 'N/A')}")
