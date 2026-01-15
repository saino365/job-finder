#!/usr/bin/env python3
"""
Read failed test cases from Registration.Login sheet.
"""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook

def get_failed_tests(sheet_name="Registration.Login", excel_path="JobFinder_TestCase.xlsx", limit=10):
    """Get failed test cases from a sheet."""
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[sheet_name]
    
    # Find Status column (usually column J, index 10)
    headers = []
    status_col_idx = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        headers.append(header)
        if header and 'Status' in str(header):
            status_col_idx = col
            break
    
    if not status_col_idx:
        print("Status column not found!")
        return []
    
    print(f"Found Status column at column {status_col_idx}")
    print(f"Headers: {headers}")
    
    # Find failed tests
    failed_tests = []
    for row_idx in range(2, sheet.max_row + 1):
        status = sheet.cell(row=row_idx, column=status_col_idx).value
        if status and 'Fail' in str(status):
            # Get all row data
            test_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    test_data[header] = cell_value
            
            test_data['row_number'] = row_idx
            failed_tests.append(test_data)
    
    print(f"\nFound {len(failed_tests)} failed test cases")
    return failed_tests[:limit]

if __name__ == "__main__":
    failed = get_failed_tests(limit=10)
    
    print("\n" + "="*60)
    print("FIRST 10 FAILED TEST CASES:")
    print("="*60)
    
    for idx, test in enumerate(failed, 1):
        print(f"\n{idx}. Row {test.get('row_number')}")
        print(f"   Test Case No: {test.get('Test Case No', 'N/A')}")
        print(f"   Test Cases: {test.get('Test Cases', 'N/A')}")
        print(f"   Test Scenario: {test.get('Test Scenario', 'N/A')}")
        print(f"   Status: {test.get('Status', 'N/A')}")
        print(f"   Defect: {test.get('Defect', 'N/A')}")
        if test.get('Actual Result'):
            print(f"   Actual Result: {str(test.get('Actual Result'))[:100]}...")
        if test.get('Expected Results'):
            print(f"   Expected: {str(test.get('Expected Results'))[:100]}...")
