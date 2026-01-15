#!/usr/bin/env python3
"""
Get next 10 failed test cases (skip already fixed ones).
"""
import sys
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import load_workbook

def get_next_failed_tests(sheet_name="Registration.Login", excel_path="JobFinder_TestCase.xlsx", skip_rows=[], limit=10):
    """Get failed test cases, skipping already fixed ones."""
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[sheet_name]
    
    # Find Status column
    headers = []
    status_col_idx = None
    fix_summary_col_idx = None
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        headers.append(header)
        if header and 'Status' in str(header):
            status_col_idx = col
        if header and 'Fix Summary' in str(header):
            fix_summary_col_idx = col
    
    if not status_col_idx:
        print("Status column not found!")
        return []
    
    # Find failed tests (Status = Failed and no Fix Summary)
    failed_tests = []
    for row_idx in range(2, sheet.max_row + 1):
        # Skip already processed rows
        if row_idx in skip_rows:
            continue
            
        status = sheet.cell(row=row_idx, column=status_col_idx).value
        fix_summary = None
        if fix_summary_col_idx:
            fix_summary = sheet.cell(row=row_idx, column=fix_summary_col_idx).value
        
        # Only get tests that are still marked as Failed and not yet fixed
        if status and 'Fail' in str(status) and not fix_summary:
            test_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    test_data[header] = cell_value
            
            test_data['row_number'] = row_idx
            failed_tests.append(test_data)
            
            if len(failed_tests) >= limit:
                break
    
    return failed_tests

if __name__ == "__main__":
    # Skip rows 4, 5, 11, 14, 15, 16, 17, 20, 21, 22 (already fixed)
    skip_rows = [4, 5, 11, 14, 15, 16, 17, 20, 21, 22]
    failed = get_next_failed_tests(skip_rows=skip_rows, limit=10)
    
    print(f"\nFound {len(failed)} more failed test cases")
    print("="*60)
    
    for idx, test in enumerate(failed, 1):
        print(f"\n{idx}. Row {test.get('row_number')}")
        print(f"   Test Case No: {test.get('Test Case No', 'N/A')}")
        print(f"   Test Cases: {test.get('Test Cases', 'N/A')}")
        print(f"   Status: {test.get('Status', 'N/A')}")
        print(f"   Defect: {test.get('Defect', 'N/A')}")
        if test.get('Expected Results'):
            print(f"   Expected: {str(test.get('Expected Results'))[:80]}...")
        if test.get('Actual Result'):
            print(f"   Actual: {str(test.get('Actual Result'))[:80]}...")
